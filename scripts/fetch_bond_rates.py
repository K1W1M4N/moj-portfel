#!/usr/bin/env python3
"""
Skrypt aktualizujący stawki obligacji skarbowych w bondRates.js
Źródło: strona marciniwuc.com (kalkulator aktualizowany co miesiąc)
Uruchamiany przez GitHub Actions 2. dnia każdego miesiąca
"""

import re
import sys
import json
import requests
from datetime import datetime
from bs4 import BeautifulSoup
import openpyxl
from io import BytesIO

BLOG_URL = "https://marciniwuc.com/obligacje-indeksowane-inflacja-kalkulator/"

# Mapowanie nazw w Excelu na klucze w bondRates.js
BOND_TYPE_MAP = {
    "TOS": "TOS",
    "COI": "COI", 
    "EDO": "EDO",
    "ROS": "ROS",
    "ROD": "ROD",
    "ROR": "ROR",
    "DOR": "DOR",
}

def get_excel_url():
    """Pobierz URL do aktualnego kalkulatora Excel ze strony bloga"""
    print("Pobieram stronę bloga...")
    resp = requests.get(BLOG_URL, headers={"User-Agent": "Mozilla/5.0"}, timeout=30)
    resp.raise_for_status()
    
    soup = BeautifulSoup(resp.text, "html.parser")
    
    # Szukaj linku do pliku XLSX kalkulatora
    for link in soup.find_all("a", href=True):
        href = link["href"]
        if "Kalkulator-obligacji" in href and href.endswith(".xlsx"):
            print(f"Znalazłem kalkulator: {href}")
            return href
    
    raise Exception("Nie znalazłem linku do kalkulatora Excel na stronie!")

def download_excel(url):
    """Pobierz plik Excel"""
    print(f"Pobieram Excel: {url}")
    resp = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=60)
    resp.raise_for_status()
    return BytesIO(resp.content)

def extract_rates_from_excel(excel_file):
    """
    Wyciągnij stawki roku 1 dla każdego typu obligacji z Excela.
    Kalkulator Marcina Iwucia ma arkusz 'WPISZ ZAŁOŻENIA' z tabelą stawek.
    """
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    print(f"Arkusze w pliku: {wb.sheetnames}")
    
    rates = {}
    now = datetime.now()
    year_month = f"{now.year}-{str(now.month).padStart(2, '0')}" if False else f"{now.year}-{now.month:02d}"
    
    # Szukaj w arkuszu z założeniami
    sheet_names_to_try = ["WPISZ ZAŁOŻENIA", "OBLIGACJE", "Sheet1"]
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\nArkusz: {sheet_name}")
        
        # Przeszukaj wszystkie komórki w poszukiwaniu nazw obligacji i stawek
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    val = cell.value.strip()
                    # Szukaj wierszy z nazwami obligacji
                    for bond_type in BOND_TYPE_MAP.keys():
                        if bond_type in val and "%" in str(ws.cell(cell.row, cell.column + 1).value or ""):
                            rate_cell = ws.cell(cell.row, cell.column + 1)
                            rate_val = rate_cell.value
                            if isinstance(rate_val, (int, float)) and 0 < rate_val < 0.3:
                                rates[bond_type] = rate_val
                                print(f"  {bond_type}: {rate_val*100:.2f}%")
                            elif isinstance(rate_val, str) and "%" in rate_val:
                                try:
                                    r = float(rate_val.replace("%", "").replace(",", ".").strip()) / 100
                                    if 0 < r < 0.3:
                                        rates[bond_type] = r
                                        print(f"  {bond_type}: {r*100:.2f}%")
                                except:
                                    pass
    
    # Jeśli nie znaleziono przez automatyczne parsowanie,
    # użyj stawek z arkusza "WPISZ ZAŁOŻENIA" wiersz z oprocentowaniem
    if not rates:
        print("\nPróba alternatywnego parsowania...")
        if "WPISZ ZAŁOŻENIA" in wb.sheetnames:
            ws = wb["WPISZ ZAŁOŻENIA"]
            for row in ws.iter_rows(values_only=True):
                row_str = [str(c) for c in row if c is not None]
                for bond_type in BOND_TYPE_MAP.keys():
                    if any(bond_type in s for s in row_str):
                        for val in row:
                            if isinstance(val, float) and 0.01 < val < 0.20:
                                rates[bond_type] = val
                                print(f"  {bond_type}: {val*100:.2f}%")
                                break
    
    return rates, year_month

def update_bond_rates_js(rates, year_month, js_file_path="src/bondRates.js"):
    """Zaktualizuj plik bondRates.js o nowe stawki"""
    
    if not rates:
        print("Brak stawek do aktualizacji!")
        return False
    
    print(f"\nAktualizuję {js_file_path} dla miesiąca {year_month}...")
    
    with open(js_file_path, "r", encoding="utf-8") as f:
        content = f.read()
    
    updated = False
    
    for bond_type, rate in rates.items():
        # Szukaj sekcji dla danego typu w BOND_RATES_HISTORY
        # Format: "YYYY-MM":0.XXXX, na końcu sekcji danego typu
        
        # Sprawdź czy ten miesiąc już istnieje
        if f'"{year_month}"' in content:
            print(f"  {bond_type}: miesiąc {year_month} już istnieje, pomijam")
            continue
        
        # Znajdź ostatni wpis w sekcji danego typu i dodaj po nim nowy
        # Szukamy wzorca: ostatni "YYYY-MM":X.XXXX w bloku danego TYP
        pattern = rf'({bond_type}:\s*{{[^}}]*?)("20\d\d-\d\d":\s*[\d.]+,?\s*\n\s*}})'
        
        new_entry = f'"  \\n    \\"{year_month}\\":{rate:.4f},'
        
        # Prostsze podejście: znajdź sekcję i ostatnią linijkę z datą
        # Wzorzec: znajdź blok TOS: { ... } i dodaj przed zamykającym }
        block_pattern = rf'({bond_type}:\s*\{{)(.*?)(\n  \}})'
        
        def add_rate(match):
            block_start = match.group(1)
            block_content = match.group(2)
            block_end = match.group(3)
            
            # Sprawdź czy już jest ten miesiąc
            if year_month in block_content:
                return match.group(0)
            
            # Dodaj nowy wpis na końcu bloku
            new_line = f'\n    "{year_month}":{rate:.4f},'
            return block_start + block_content + new_line + block_end
        
        new_content = re.sub(block_pattern, add_rate, content, flags=re.DOTALL)
        
        if new_content != content:
            content = new_content
            print(f"  ✓ {bond_type}: dodano {year_month} = {rate*100:.2f}%")
            updated = True
        else:
            print(f"  ✗ {bond_type}: nie udało się zaktualizować")
    
    if updated:
        with open(js_file_path, "w", encoding="utf-8") as f:
            f.write(content)
        print(f"\n✓ Zapisano {js_file_path}")
    
    return updated

def main():
    print("=== Aktualizacja stawek obligacji skarbowych ===")
    print(f"Data: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n")
    
    try:
        # 1. Pobierz URL do aktualnego Excela
        excel_url = get_excel_url()
        
        # 2. Pobierz Excel
        excel_file = download_excel(excel_url)
        
        # 3. Wyciągnij stawki
        rates, year_month = extract_rates_from_excel(excel_file)
        print(f"\nZnalezione stawki dla {year_month}: {rates}")
        
        # 4. Zaktualizuj bondRates.js
        updated = update_bond_rates_js(rates, year_month)
        
        if updated:
            print("\n✓ Aktualizacja zakończona sukcesem!")
        else:
            print("\n! Brak zmian do zapisania")
            
    except Exception as e:
        print(f"\n✗ Błąd: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()
